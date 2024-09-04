
from openpyxl import Workbook, load_workbook

# Excel 파일 불러오기
load_wb = load_workbook("C:\\Users\\WD\\Downloads\\고장상황_20240904_110013.xlsx", data_only=True)

# 시트 이름으로 불러오기
load_ws = load_wb['Sheet1']

# 셀 주소로 값 출력
print(load_ws['B2'].value)

# 셀 좌표로 값 출력 (예: (2, 2)인 셀의 값 출력)
print(load_ws.cell(row=2, column=2).value)
